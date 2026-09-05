#!/usr/bin/env python3
"""data_local の xlsx から _data の CSV を作り直す。

    python3 scripts/from_xlsx.py [xlsx のパス]

既定の入力は data_local/takahashi_publications_awards.xlsx。
シートは Publications / Awards / Talks。列はそのまま CSV になる。
sortkey は xlsx に持たせず、ここで YYYYMMDD に計算し直して並べ替える
（build_cv.py の pub_sort_key と同じ規則）。

xlsx は data_local/ にあり gitignore 済み。手元でだけ動くスクリプト。
Google スプレッドシート同期に切り替えた場合は scripts/sync_sheet.py が
同じことをするので、こちらは使わなくてよい。
"""
import csv, sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from build_cv import PUB_COLS, pub_sort_key   # noqa: E402

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl が要る。pip3 install --user openpyxl")

ROOT = Path(__file__).resolve().parent.parent
XLSX = Path(sys.argv[1]) if len(sys.argv) > 1 else ROOT / "data_local/takahashi_publications_awards.xlsx"
OUT = ROOT / "_data"
SHEETS = {"Publications": "publications.csv", "Awards": "awards.csv", "Talks": "talks.csv"}


def cell(v):
    """数値として読まれた year / month / day の先頭 0 が落ちないようにする。"""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    return str(v).strip()


def main():
    if not XLSX.exists():
        sys.exit(f"{XLSX} が無い")
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    for sheet, fname in SHEETS.items():
        if sheet not in wb.sheetnames:
            sys.exit(f"シート「{sheet}」が無い")
        ws = wb[sheet]
        header = [cell(c.value) for c in ws[1]]
        rows = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            d = {h: cell(v) for h, v in zip(header, r)}
            if not any(d.values()):          # 空行は捨てる
                continue
            for k in ("year", "month", "day"):
                if k in d and d[k]:
                    d[k] = d[k].zfill(2) if k != "year" else d[k]
            rows.append(d)

        if fname == "awards.csv":
            cols = header
            rows.sort(key=lambda x: (x.get("year", ""), x.get("month", "")), reverse=True)
        else:
            cols = PUB_COLS
            for d in rows:
                d["sortkey"] = f"{d.get('year','')}{d.get('month','')}{(d.get('day') or '').zfill(2) or '00'}"
            rows.sort(key=pub_sort_key)

        with (OUT / fname).open("w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=cols, quoting=csv.QUOTE_ALL,
                               extrasaction="ignore", restval="")
            w.writeheader()
            w.writerows(rows)
        print(f"{fname}: {len(rows)} 件")


if __name__ == "__main__":
    main()
