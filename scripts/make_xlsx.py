#!/usr/bin/env python3
"""_data の CSV から、高橋亮個人の publication / award / talk をまとめた xlsx を作る。

出力先: my_web/data_local/takahashi_publications_awards.xlsx
data_local/ と *.xlsx は .gitignore 済み（＝リポジトリには入らない）。
Google スプレッドシートに取り込むときは、このファイルをそのままアップロードすればよい。
"""
import csv
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / "_data"
OUTDIR = ROOT / "data_local"
OUTDIR.mkdir(exist_ok=True)
OUT = OUTDIR / "takahashi_publications_awards.xlsx"

SHEETS = [
    ("Publications", "publications.csv"),
    ("Awards", "awards.csv"),
    ("Talks", "talks.csv"),
]

wb = openpyxl.Workbook()
wb.remove(wb.active)

for title, fname in SHEETS:
    path = DATA / fname
    if not path.exists():
        continue
    rows = list(csv.reader(path.open(encoding="utf-8")))
    ws = wb.create_sheet(title)
    for r in rows:
        ws.append(r)
    # 見出し行を固定して読みやすくする
    for c in range(1, ws.max_column + 1):
        ws.cell(row=1, column=c).font = Font(bold=True)
        ws.cell(row=1, column=c).alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    widths = {}
    for row in ws.iter_rows(values_only=True):
        for i, v in enumerate(row, start=1):
            widths[i] = max(widths.get(i, 10), min(len(str(v or "")) + 2, 60))
    for i, w in widths.items():
        ws.column_dimensions[get_column_letter(i)].width = w
    print(f"{title}: {ws.max_row - 1} 件")

wb.save(OUT)
print("saved:", OUT)
